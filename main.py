"""Main entry point for txt2sql benchmark suite.

Loads test cases, runs metrics, and exports Excel report.

Usage:
    python main.py --input data/test_cases.json --output results/report.xlsx --weight 0.3
"""

import json
import sys
import argparse
from datetime import datetime
from pathlib import Path
from typing import List

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from model import TestCase, MetricResult
from config import (
    WEIGHT_TABLE_SIM,
    WEIGHT_SEMANTIC_SIM,
    WEIGHT_LLM_SCORE,
    WEIGHT_VES,
    LM_STUDIO_API_URL,
    EMBEDDING_MODEL,
    SQLITE_DB_PATH,
    TABLE_SIM_ORDER_SENSITIVE,
    TABLE_SIM_ORDER_MISMATCH_WEIGHT,
    COLUMN_SELECTION_LLM_ENABLED,
    COLUMN_SELECTION_MODEL,
    LLM_JUDGE_MODEL,
)
from mock_database import MockDatabaseExecutor
from metric import run_benchmark
from report import generate_html_report


def load_test_cases(json_file: str) -> List[TestCase]:
    """Load test cases from JSON file.

    Expected format:
    [
        {
            "natural_language": "...",
            "generated_sql": "...",
            "expected_sql": "..."
        },
        ...
    ]
    """
    with open(json_file, "r") as f:
        data = json.load(f)

    test_cases = []
    for item in data:
        test_cases.append(
            TestCase(
                natural_language=item.get("natural_language", ""),
                generated_sql=item.get("generated_sql", ""),
                expected_sql=item.get("expected_sql", ""),
            )
        )

    return test_cases


def export_to_excel(
    results: List[MetricResult],
    summary_stats: dict,
    output_file: str,
) -> None:
    """Export benchmark results to Excel with multiple sheets.

    Creates sheets:
    1. Summary: Overall statistics
    2. Results: Per-query metrics
    3. Info: Configuration and instructions
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Summary Statistics
    ws_summary = wb.create_sheet("Summary")
    _populate_summary_sheet(ws_summary, summary_stats, results)

    # Sheet 2: Per-Query Results
    ws_results = wb.create_sheet("Results")
    _populate_results_sheet(ws_results, results)

    # Sheet 3: Info
    ws_info = wb.create_sheet("Info")
    _populate_info_sheet(ws_info, summary_stats)

    # Sheet 4: Interactive composite score dashboard
    ws_dashboard = wb.create_sheet("Dashboard")
    _populate_dashboard_sheet(ws_dashboard, results, summary_stats)

    # Save workbook
    wb.save(output_file)
    print(f"\n✓ Excel report saved to: {output_file}")


def _populate_dashboard_sheet(
    ws, results: List[MetricResult], summary_stats: dict
) -> None:
    """Populate an interactive Dashboard with editable weight cells and a live chart.

    Layout:
    - Rows 1-2:  Title and instructions
    - Rows 4-9:  Weight input table (Key | Component | Weight [yellow, editable])
    - Row 11:    Data table header
    - Row 12+:   Per-query scores
                   Cols C-F: raw scores (static)
                   Cols G-J: Excel formulas = weight_cell * raw_score
                   Col  K:   Composite = G+H+I+J
    - Below data: Stacked bar chart referencing cols G-J
                  (updates automatically when weight cells change)
    """
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    blue_header = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    dark_blue = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    light_blue = PatternFill(
        start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
    )

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    ws["A1"].value = "Interactive Composite Score Dashboard"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = dark_blue
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    ws["A2"].value = (
        "Edit the yellow Weight cells (column C, rows 5\u20138).  "
        "The chart and Composite column recalculate automatically."
    )
    ws["A2"].font = Font(italic=True, size=10, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Weight input table ────────────────────────────────────────────────────
    for col, label in [
        (1, "Key"),
        (2, "Component"),
        (3, "Weight  \u270f  (edit me)"),
    ]:
        c = ws.cell(row=4, column=col, value=label)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = blue_header
        c.alignment = Alignment(horizontal="center")
        c.border = border

    weight_inputs = [
        (5, "W1", "Table Similarity (S_T)", summary_stats["w1"]),
        (6, "W2", "Semantic Similarity (S_C)", summary_stats["w2"]),
        (7, "W3", "LLM Score", summary_stats["w3"]),
        (8, "W4", "VES", summary_stats["w4"]),
    ]
    for row_num, key, label, weight in weight_inputs:
        c_key = ws.cell(row=row_num, column=1, value=key)
        c_key.font = Font(bold=True)
        c_key.border = border
        c_key.alignment = Alignment(horizontal="center")

        ws.cell(row=row_num, column=2, value=label).border = border

        c_w = ws.cell(row=row_num, column=3, value=round(weight, 4))
        c_w.fill = yellow
        c_w.border = border
        c_w.alignment = Alignment(horizontal="center")
        c_w.number_format = "0.00"

    # Sum row
    c_sum_lbl = ws.cell(row=9, column=2, value="Sum (should equal 1.0)")
    c_sum_lbl.font = Font(italic=True)
    c_sum_lbl.border = border
    c_sum = ws.cell(row=9, column=3, value="=C5+C6+C7+C8")
    c_sum.border = border
    c_sum.alignment = Alignment(horizontal="center")
    c_sum.number_format = "0.00"

    # ── Data table ────────────────────────────────────────────────────────────
    TABLE_HDR = 11
    DATA_ROW = 12
    # Weight cell references — changing these cells updates all formulas + chart
    W_REFS = ["$C$5", "$C$6", "$C$7", "$C$8"]

    tbl_headers = [
        "Query",
        "Natural Language",
        "S_T",
        "S_C",
        "LLM",
        "VES",
        "W1\u00b7S_T",
        "W2\u00b7S_C",
        "W3\u00b7LLM",
        "W4\u00b7VES",
        "Composite",
    ]
    for col_idx, header in enumerate(tbl_headers, start=1):
        c = ws.cell(row=TABLE_HDR, column=col_idx, value=header)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = blue_header
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for i, result in enumerate(results):
        r = DATA_ROW + i
        ws.cell(row=r, column=1, value=f"Q{i + 1}").alignment = Alignment(
            horizontal="center"
        )
        ws.cell(row=r, column=2, value=result.test_case.natural_language)

        raw_vals = [
            result.table_sim,
            result.semantic_sim,
            result.llm_score,
            result.ves,
        ]
        for j, val in enumerate(raw_vals):
            c = ws.cell(row=r, column=3 + j, value=round(val, 4))
            c.alignment = Alignment(horizontal="center")
            c.number_format = "0.0000"

        # Formula cells: weight_ref * raw_score_cell
        for j in range(4):
            raw_col_letter = get_column_letter(3 + j)
            c = ws.cell(
                row=r,
                column=7 + j,
                value=f"={W_REFS[j]}*{raw_col_letter}{r}",
            )
            c.fill = light_blue
            c.alignment = Alignment(horizontal="center")
            c.number_format = "0.0000"

        c_comp = ws.cell(row=r, column=11, value=f"=G{r}+H{r}+I{r}+J{r}")
        c_comp.fill = light_blue
        c_comp.alignment = Alignment(horizontal="center")
        c_comp.number_format = "0.0000"

        for col in range(1, 12):
            ws.cell(row=r, column=col).border = border

    last_data_row = DATA_ROW + len(results) - 1

    # ── Stacked bar chart (references formula cells → auto-updates with weights) ──
    if results:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = "Composite Score Breakdown by Query"
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Query"
        chart.height = 14
        chart.width = 24

        # Weighted-component columns G-J (cols 7-10); header row provides series names
        data = Reference(
            ws, min_col=7, max_col=10, min_row=TABLE_HDR, max_row=last_data_row
        )
        chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=DATA_ROW, max_row=last_data_row)
        chart.set_categories(cats)
        chart.legend.position = "b"
        ws.add_chart(chart, f"A{last_data_row + 3}")

    # ── Column widths & freeze ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 16
    for col_letter in ["D", "E", "F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col_letter].width = 12
    ws.freeze_panes = f"A{DATA_ROW}"


def _populate_summary_sheet(
    ws, summary_stats: dict, results: List[MetricResult]
) -> None:
    """Populate summary statistics sheet."""
    # Header
    ws["A1"] = "Benchmark Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws.merge_cells("A1:B1")

    # Metadata
    ws["A3"] = "Report Generated"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws["A4"] = "Total Tests"
    ws["B4"] = int(summary_stats["total_tests"])

    ws["A5"] = "Weights"
    ws["B5"] = (
        f"W1(S_T)={summary_stats['w1']:.2f}  "
        f"W2(S_C)={summary_stats['w2']:.2f}  "
        f"W3(LLM)={summary_stats['w3']:.2f}  "
        f"W4(VES)={summary_stats['w4']:.2f}"
    )

    # Metrics
    ws["A7"] = "Metric"
    ws["B7"] = "Value"
    for cell in ["A7", "B7"]:
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )

    row = 8
    metrics = [
        ("Avg Semantic Similarity (S_C)", f"{summary_stats['avg_semantic_sim']:.4f}"),
        ("Avg Table Similarity (S_T)", f"{summary_stats['avg_table_sim']:.4f}"),
        ("Avg LLM Score", f"{summary_stats['avg_llm_score']:.4f}"),
        ("Avg VES", f"{summary_stats['avg_ves']:.4f}"),
        ("Avg Composite Score", f"{summary_stats['avg_composite_score']:.4f}"),
        ("Total Time (ms)", f"{summary_stats['total_time_ms']:.1f}"),
    ]

    for metric_name, metric_value in metrics:
        ws[f"A{row}"] = metric_name
        ws[f"B{row}"] = metric_value
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def _populate_results_sheet(ws, results: List[MetricResult]) -> None:
    """Populate per-query results sheet."""
    # Header
    headers = [
        "Query #",
        "Natural Language",
        "Generated SQL",
        "Expected SQL",
        "Semantic Sim (S_C)",
        "Table Sim (S_T)",
        "LLM Score",
        "VES",
        "Composite Score",
        "Eval Gen Columns",
        "Eval Exp Columns",
        "Column Judge Source",
        "Column Judge Confidence",
        "Gen Exec Time (ms)",
        "Ref Exec Time (ms)",
        "Total Time (ms)",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # Data rows
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1).value = row_idx - 1
        ws.cell(row=row_idx, column=2).value = result.test_case.natural_language
        ws.cell(row=row_idx, column=3).value = result.test_case.generated_sql
        ws.cell(row=row_idx, column=4).value = result.test_case.expected_sql
        ws.cell(row=row_idx, column=5).value = f"{result.semantic_sim:.4f}"
        ws.cell(row=row_idx, column=6).value = f"{result.table_sim:.4f}"
        ws.cell(row=row_idx, column=7).value = f"{result.llm_score:.4f}"
        ws.cell(row=row_idx, column=8).value = f"{result.ves:.4f}"
        ws.cell(row=row_idx, column=9).value = f"{result.composite_score:.4f}"
        ws.cell(row=row_idx, column=10).value = ", ".join(
            result.selected_generated_columns
        )
        ws.cell(row=row_idx, column=11).value = ", ".join(
            result.selected_expected_columns
        )
        ws.cell(row=row_idx, column=12).value = result.column_selection_source
        ws.cell(row=row_idx, column=13).value = (
            f"{result.column_selection_confidence:.2f}"
        )
        ws.cell(row=row_idx, column=14).value = f"{result.execution_time_gen_ms:.1f}"
        ws.cell(row=row_idx, column=15).value = f"{result.execution_time_ref_ms:.1f}"
        ws.cell(row=row_idx, column=16).value = f"{result.execution_time_ms:.1f}"

        # Center align numeric columns
        for col in [1, 5, 6, 7, 8, 9, 13, 14, 15, 16]:
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 24
    ws.column_dimensions["K"].width = 24
    ws.column_dimensions["L"].width = 18
    ws.column_dimensions["M"].width = 18
    ws.column_dimensions["N"].width = 16
    ws.column_dimensions["O"].width = 16
    ws.column_dimensions["P"].width = 14


def _populate_info_sheet(ws, summary_stats: dict) -> None:
    """Populate information sheet."""
    ws["A1"] = "Configuration & Instructions"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "Current Settings"
    ws["A3"].font = Font(bold=True, size=11)

    ws["A4"] = "Weights"
    ws["B4"] = (
        f"W1(S_T)={summary_stats['w1']:.2f}  "
        f"W2(S_C)={summary_stats['w2']:.2f}  "
        f"W3(LLM)={summary_stats['w3']:.2f}  "
        f"W4(VES)={summary_stats['w4']:.2f}"
    )

    ws["A5"] = "Formula"
    ws["B5"] = "Composite = (W1*S_T) + (W2*S_C) + (W3*LLM_SCORE) + (W4*VES)"

    ws["A6"] = "Intent-Aware Columns"
    ws["B6"] = f"enabled={COLUMN_SELECTION_LLM_ENABLED}, model={COLUMN_SELECTION_MODEL}"

    ws["A7"] = "LLM Judge Model"
    ws["B7"] = LLM_JUDGE_MODEL

    ws["A9"] = "Instructions for Different Weights"
    ws["A9"].font = Font(bold=True, size=11)

    instructions = [
        "To test with different weights, re-run the benchmark with --w1 --w2 --w3 --w4:",
        "",
        "Examples:",
        "  python main.py --w1 0.3 --w2 0.2 --w3 0.3 --w4 0.2   # Default",
        "  python main.py --w1 0.4 --w2 0.2 --w3 0.2 --w4 0.2   # Emphasize table sim",
        "  python main.py --w1 0.25 --w2 0.25 --w3 0.25 --w4 0.25  # Equal weights",
        "",
        "Weights should sum to 1.0 for a normalized composite score.",
        "W1 = Table Similarity (S_T): result-set correctness",
        "W2 = Semantic Similarity (S_C): SQL intent similarity via embeddings",
        "W3 = LLM Score: LLM-as-judge evaluation of query quality",
        "W4 = VES: Valid Efficiency Score (execution speed vs reference)",
    ]

    row = 10
    for instruction in instructions:
        ws[f"A{row}"] = instruction
        row += 1

    ws["A22"] = "Metric Definitions"
    ws["A22"].font = Font(bold=True, size=11)

    definitions = [
        ("EM (Exact Match)", "Binary: 1 if normalized SQL strings match, 0 otherwise"),
        ("EX (Execution Accuracy)", "Binary: 1 if query results match, 0 otherwise"),
        ("S_C (Semantic Similarity)", "Cosine similarity of SQL embeddings [0,1]"),
        ("S_T (Table Similarity)", "Edit distance-based column similarity [0,1]"),
        ("LLM Score", "LLM-as-judge evaluation of SQL quality [0,1]"),
        (
            "VES (Valid Efficiency Score)",
            "sqrt(ref_time / gen_time), capped at 1.0, 0 if EX=False",
        ),
        (
            "Composite Score",
            "Weighted combination: (W1*S_T) + (W2*S_C) + (W3*LLM) + (W4*VES)",
        ),
    ]

    row = 23
    for name, definition in definitions:
        ws[f"A{row}"] = name
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = definition
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 60


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="txt2sql Benchmark Suite",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python main.py
    python main.py --input data/sakila_test_cases.json --output results/report.xlsx
  python main.py --weight 0.5
        """,
    )

    parser.add_argument(
        "--input",
        type=str,
        default="data/sakila_test_cases.json",
        help="Input JSON file with test cases (default: data/sakila_test_cases.json)",
    )
    parser.add_argument(
        "--output",
        type=str,
        default="results/benchmark_report.xlsx",
        help="Output Excel file for report (default: results/benchmark_report.xlsx)",
    )
    parser.add_argument(
        "--w1",
        type=float,
        default=WEIGHT_TABLE_SIM,
        help=f"Weight for Table Similarity S_T (default: {WEIGHT_TABLE_SIM})",
    )
    parser.add_argument(
        "--w2",
        type=float,
        default=WEIGHT_SEMANTIC_SIM,
        help=f"Weight for Semantic Similarity S_C (default: {WEIGHT_SEMANTIC_SIM})",
    )
    parser.add_argument(
        "--w3",
        type=float,
        default=WEIGHT_LLM_SCORE,
        help=f"Weight for LLM Score (default: {WEIGHT_LLM_SCORE})",
    )
    parser.add_argument(
        "--w4",
        type=float,
        default=WEIGHT_VES,
        help=f"Weight for VES (default: {WEIGHT_VES})",
    )
    parser.add_argument(
        "--table-order-sensitive",
        action="store_true",
        default=TABLE_SIM_ORDER_SENSITIVE,
        help=("Use order-sensitive table similarity (default is order-insensitive)."),
    )
    parser.add_argument(
        "--table-order-mismatch-weight",
        type=float,
        default=TABLE_SIM_ORDER_MISMATCH_WEIGHT,
        help=(
            "Soft penalty weight for shuffled row order in order-insensitive "
            "mode (default: 0.0)."
        ),
    )

    args = parser.parse_args()

    print("\n" + "=" * 70)
    print("txt2sql Benchmark Suite")
    print("=" * 70)

    # Check input file exists
    if not Path(args.input).exists():
        print(f"❌ Error: Input file not found: {args.input}")
        return

    # Create output directory
    output_dir = Path(args.output).parent
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load test cases
    print(f"\n📂 Loading test cases from: {args.input}")
    test_cases = load_test_cases(args.input)
    print(f"   ✓ Loaded {len(test_cases)} test case(s)")

    # Initialize services
    print(f"\n🔌 Initializing services...")

    # Initialize OpenAI client pointing to LM Studio
    try:
        openai_client = OpenAI(api_key="dummy", base_url=LM_STUDIO_API_URL)
        print(f"   ✓ Connected to LM Studio at {LM_STUDIO_API_URL}")
        print(f"   ✓ Embedding model: {EMBEDDING_MODEL}")
        if COLUMN_SELECTION_LLM_ENABLED:
            print(f"   ✓ Column selection model: {COLUMN_SELECTION_MODEL}")
        print(f"   ✓ LLM judge model: {LLM_JUDGE_MODEL}")
    except Exception as e:
        print(f"   ⚠️  Warning: Could not connect to LM Studio: {e}")
        print(f"      Make sure LM Studio is running at {LM_STUDIO_API_URL}")
        print(
            f"      Continuing with mock embeddings (all similarities will be random)..."
        )
        openai_client = None

    # Initialize SQLite executor (sakila.db)
    db_executor = MockDatabaseExecutor(SQLITE_DB_PATH)
    print(f"   ✓ SQLite database initialized: {SQLITE_DB_PATH}")

    # Run benchmark
    print(f"\n📊 Running benchmark...")
    print(f"   Composite = (W1*S_T) + (W2*S_C) + (W3*LLM) + (W4*VES)")
    print(
        f"   W1(S_T)={args.w1:.2f}  W2(S_C)={args.w2:.2f}  W3(LLM)={args.w3:.2f}  W4(VES)={args.w4:.2f}"
    )
    print(
        "   (Table similarity mode: "
        + ("order-sensitive" if args.table_order_sensitive else "order-insensitive")
        + ")"
    )
    print(f"   (Order mismatch weight: {args.table_order_mismatch_weight:.2f})")
    print()

    results, summary_stats = run_benchmark(
        test_cases,
        db_executor,
        openai_client,
        w1=args.w1,
        w2=args.w2,
        w3=args.w3,
        w4=args.w4,
        table_order_sensitive=args.table_order_sensitive,
        table_order_mismatch_weight=args.table_order_mismatch_weight,
    )

    # Print summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"Total Tests:        {int(summary_stats['total_tests'])}")
    print(f"Avg Semantic Sim:   {summary_stats['avg_semantic_sim']:.4f}")
    print(f"Avg Table Sim:      {summary_stats['avg_table_sim']:.4f}")
    print(f"Avg LLM Score:      {summary_stats['avg_llm_score']:.4f}")
    print(f"Avg VES:            {summary_stats['avg_ves']:.4f}")
    print(f"Avg Composite:      {summary_stats['avg_composite_score']:.4f}")
    print(f"Total Time:         {summary_stats['total_time_ms']:.1f} ms")

    # Export to Excel
    print(f"\n📝 Exporting results to Excel...")
    export_to_excel(results, summary_stats, args.output)

    # Export interactive HTML report
    html_output = str(Path(args.output).with_suffix(".html"))
    generate_html_report(results, summary_stats, html_output)
    print(f"✓ Interactive HTML report saved to: {html_output}")

    print("\n" + "=" * 70)
    print("✓ Benchmark completed successfully!")
    print("=" * 70)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Benchmark interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\n\n❌ Error: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)
